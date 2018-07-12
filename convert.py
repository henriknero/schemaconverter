from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime, timedelta
import csv


def l_o(letter, offset):    # letter offset
    if letter == 'Z' and offset > 0:
        return 'AB'
    return chr(ord(letter)+offset)


m_to_n = {'jun': 6, 'aug': 8, 'maj': 5, 'jul': 7}       # Month to number
NAME = 'Henrik'
wb = load_workbook('schema.xlsx')
output_sheet = Workbook().create_sheet()
output_sheet['A1'] = 'Subject'
output_sheet['B1'] = 'Start Date'
output_sheet['C1'] = 'Start Time'
output_sheet['D1'] = 'End Date'
output_sheet['E1'] = 'End Time'
out_row = 2
for vecka in wb:
    row_index = 0
    for x in vecka['a']:
        if x.value == NAME:
            row_index = x.row
    # Flyttar från schema format till csv format
    if row_index != 0:

        col = 'B'
        for _ in range(5):
            if vecka[col+str(row_index)].value is not None:
                output_sheet['A' + str(out_row)] = 'Jobb'
                output_sheet['B' + str(out_row)] = datetime(2018, m_to_n[vecka[l_o(col, 2)+'1'].value],
                                                            vecka[col+'1'].value).strftime('%Y-%m-%d')  # Start date
                output_sheet['C' + str(out_row)] = vecka[col+str(row_index)].value  # Start time
                output_sheet['D' + str(out_row)] = datetime(2018, m_to_n[vecka[l_o(col,2)+'1'].value],
                                                            vecka[col+'1'].value).strftime('%Y-%m-%d')  # End date
                output_sheet['E' + str(out_row)] = vecka[l_o(col,2)+str(row_index)].value   # End time

                if '1900-01-01 00:15:00' == str(vecka[l_o(col,2) + str(row_index)].value):
                    output_sheet['D' + str(out_row)] = (datetime(2018, m_to_n[vecka[l_o(col,2)+'1'].value], vecka[col+'1'].value) + timedelta(days=1)).strftime('%Y-%m-%d')       #End date
                    output_sheet['E' + str(out_row)] = '00:15:00'
                out_row += 1
            col = l_o(col, 3)
        for _ in range(2):
            start_date = end_date = 0
            if vecka[col+str(row_index)].value is not None:
                output_sheet['A' + str(out_row)] = 'Jobb'
                start_date = datetime(2018, m_to_n[vecka[l_o(col, 3) + '1'].value],
                                        vecka[l_o(col,2) + '1'].value).strftime('%Y-%m-%d')  # Start date
                output_sheet['B' + str(out_row)] = start_date
                output_sheet['C' + str(out_row)] = vecka[col + str(row_index)].value  # Start time
                end_date = datetime(2018, m_to_n[vecka[l_o(col, 3) + '1'].value],
                                    vecka[l_o(col, 2) + '1'].value).strftime('%Y-%m-%d')  # End date
                output_sheet['D' + str(out_row)] = end_date
                output_sheet['E' + str(out_row)] = vecka[l_o(col, 2) + str(row_index)].value
                out_row += 1
            col = l_o(col, 3)
            if vecka[col + str(row_index)].value is not None:
                output_sheet['A' + str(out_row)] = 'Jobb'
                start_date = datetime(2018, m_to_n[vecka[col + '1'].value],
                                      vecka[l_o(col, -1) + '1'].value).strftime('%Y-%m-%d')  # Start date
                output_sheet['B' + str(out_row)] = start_date
                output_sheet['C' + str(out_row)] = vecka[col + str(row_index)].value  # Start time
                output_sheet['D' + str(out_row)] = start_date
                output_sheet['E' + str(out_row)] = vecka[l_o(col, 2) + str(row_index)].value
                if '1900-01-01 00:15:00' == str(vecka[l_o(col,2) + str(row_index)].value):
                    output_sheet['D' + str(out_row)] = (datetime.strptime(start_date,'%Y-%m-%d') + timedelta(days=1)).strftime('%Y-%m-%d')       #End date
                    output_sheet['E' + str(out_row)] = '00:15:00'
                out_row += 1
            col = l_o(col, 3)
    with open('out.csv', 'w') as f:
        c = csv.writer(f)
        for r in output_sheet.rows:
            c.writerow([cell.value for cell in r])
